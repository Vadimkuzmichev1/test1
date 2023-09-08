import json
from openpyxl import Workbook
from datetime import datetime


# Create a new Excel workbook
workbook = Workbook()

# Create a new sheet
sheet = workbook.active

# Get the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Create the filename with the current date and time
filename = f"result\lo_results_{current_datetime}.xlsx"

networks = {

}

contracts = {
    "arb_stargate_send": "0x53bf833a5d6c4dda888f69c22c88c9f356a41614",  # other_addr one address to all tokens
    "arb_stargate_received_usdc": "0x892785f33cdee22a30aef750f285e18c18040c3e",  # other_addr and receives: from_addr
    "arb_stargate_received_usdt": "0xb6cfcf89a7b22988bfc96632ac2a9d6dab60d641",  # other_addr and receives: from_addr 0x177d36dbe2271a4ddb2ad8304d82628eb921d790
    "arb_stargate_stake_stg": "0xfbd849e6007f9bc3cc2d6eb159c045b8dc660268",  # token id 0x6694340fc020c5e6b96567843da2df01b2ce1eb6
    "arb_testnetbridge_send": "0x0a9f824c05a74f577a536a8a0c673183a872dff4",

    "testnet_bridge_received": "",

    "avax_stargate_send": "0x45a01e4e04f14f7a4a6702c74187c5f6222033cd",  # other_addr one address to all tokens
    "avax_stargate_received_usdc": "0x1205f31718499dbf1fca446663b532ef87481fe1",  # other_addr and receives: from_addr
    "avax_stargate_received_usdt": "0x29e38769f23701a2e4a8ef0492e19da4604be62c",  # other_addr and receives: from_addr
    "avax_stargate_stake_stg": "0xca0f57d295bbce554da2c07b005b7d6565a58fce",
    "avax_btc_bridge_send": "0x2297aebd383787a160dd0d9f71508148769342e3",
    "avax_swimmer_send": "0xd12222329364c5fe5a5f87cd67e54accb2547be1",

    "swimmer_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and 0xc1a1f40d558a3e82c3981189f61ef21e17d6eb48 token id

    "matic_stargate_send": "0x45a01e4e04f14f7a4a6702c74187c5f6222033cd",  # other_addr one address to all tokens
    "matic_stargate_received_usdc": "0x1205f31718499dbf1fca446663b532ef87481fe1",  # #other_addr and receives: from_addr
    "matic_stargate_received_usdt": "0x75dc8e5f50c8221a82ca6af64af811caa983b65f",  # #other_addr and receives: from_addr
    "matic_stargate_stake_stg": "0x3ab2da31bbd886a7edf68a6b60d3cde657d3a15d",
    "matic_btc_bridge_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # tx: from addr and 0x2297aebd383787a160dd0d9f71508148769342e3 token id
    "matic_angle_send": "0x0c1ebbb61374da1a8c57cb6681bf27178360d36f",
    "celo_angle_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and  token id 0xc16b81af351ba9e64c1a069e3ab18c244a1e3049
    "celo_angle_send": "0xf1ddcaca7d17f8030ab2eb54f2d9811365efe123",
    "gnosis_angle_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and  token id 0x4b1e2c2762667331bc91648052f646d1b0d35984
    "matic_aptos_send": "0x488863d609f3a673875a914fbee7508a1de45ec6",

    "bsc_stargate_send": "0x4a364f8c717caad9a442737eb7b8a55cc6cf18d8",
    "bsc_stargate_received_usdt": "0x9aa83081aa06af7208dcc7a4cb72c94d057d2cda",
    "bsc_harmony_send": "0x0551ca9e33bada0355dfce34685ad3b73cf3ad70",  # проверить
    "bsc_core_send": "0x52e75d318cfb31f9a2edfa2dfee26b161255b233",  # проверить
    "bsc_core_received": "0x52e75d318cfb31f9a2edfa2dfee26b161255b233",
    "bsc_btc_bridge_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and 0x2297aebd383787a160dd0d9f71508148769342e3 token id
    "bsc_aptos_send": "0x2762409baa1804d94d8c0bcff8400b78bf915d5b",
    "bsc_omnisea_send": "0x5d2a31821e48c31be1c8ea54ba27ba288dbbfe46",

    "moonbeam_omnisea_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and  token id 0x16b26b9328b1b64e4aad6326c7cb94b2e8a96b4e
    "moonbeam_omnisea_send": "0xa82c3c40b3386b8725c98f48ed420cac523d7a52",

    "aptos_received": "",

    "ftm_stargate_send": "0xaf5191b0de278c7286d6c7cc6ab6bb8a73ba2cd6",
    "ftm_stargate_received_usdc": "0x12edea9cd262006cc3c4e77c90d2cd2dd4b1eb97",
    "ftm_omnisea_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and  token id 0xc72633f995e98ac3bb8a89e6a9c4af335c3d6e44

    "op_stargate_send": "0xb0d502e938ed5f4df2e681fe6e419ff29631d62b",  # 0x7F5c764cBc14f9669B88837ca1490cCa17c31607
    "op_stargate_received_usdc1": "0x81e792e5a9003cc1c8bf5569a00f34b65d75b017",
    "op_stargate_received_usdc2": "0xdecc0c09c3b5f6e92ef4184125d5648a66e35298",#

    "core_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # tx: from addr
    "core_send": "0xa4218e1f39da4aadac971066458db56e901bcbde",  # other_addr

    "harmony_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # tx: from addr
    "harmony_omnisea_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and  token id 0x20035f39c6c6cb6c0d41bb88d8f443848389b809

    "metis_omnisea_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and  token id 0x5b19bd330a84c049b62d5b0fc2ba120217a18c1c

    "klay_kingdom_send": "0x6d5b86eac9097ea4a94b2b69cd4854678b89f839",
    "klay_kingdom_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and  token id  0xe7a1b580942148451e47b92e95aeb8d31b0aca37

    "dfk_kingdom_send": "0x501cdc4ef10b63219704bf6adb785dfccb06dee2",
    "dfk_kingdom_received": "0xe93685f3bba03016f02bd1828badd6195988d950",  # from addr and  token id 0x576c260513204392f0ec0bc865450872025cb1ca

    "ftm_abracadabra_send": "0xc5c01568a3b5d8c203964049615401aaf0783191",
    "movr_abracadabra_received": "0xe93685f3bba03016f02bd1828badd6195988d950", #  # from addr and  token id 0x0cae51e1032e8461f4806e26332c030e34de3adb

}

with open("wallets.txt", "r") as f:
    wallets = [row.strip() for row in f if row.strip()]

wallet_temp = {
    "total_usd": 0,
    "total_trans": 0,
    'nft_trans': 0,
    "first_day": "",
    'last_day': "",
    'days': 0,
    'weeks': 0,
    'months': 0,
    ''

    "arb_in": 0,
    "arb_from": 0,

    "testnet_in": 0,
    "testnet_from": 0,

    "avax_in": 0,
    "avax_from": 0,

    "matic_in": 0,
    "matic_from": 0,

    "bsc_in": 0,
    "bsc_from": 0,

    "ftm_in": 0,
    "ftm_from": 0,

    "op_in": 0,
    "op_from": 0,

    "core_in": 0,
    "core_from": 0,

    "hmy_in": 0,
    "hmy_from": 0,

    "celo_in": 0,
    "celo_from": 0,

    "gnosis_in": 0,
    "gnosis_from": 0,

    "aptos_in": 0,
    "aptos_from": 0,

    "swm_in": 0,
    "swm_from": 0,

    "moonbeam_in": 0,
    "moonbeam_from": 0,
    "metis_in": 0,

    "klay_in": 0,
    "klay_from": 0,

    "dfk_in": 0,
    "dfk_from": 0,

    "arb_stargate_send": 0,
    "arb_stargate_received_usdc": 0,
    "arb_stargate_received_usdt": 0,
    "arb_stargate_stake_stg": 0,
    "arb_testnetbridge_send": 0,

    "testnet_bridge_received": 0,

    "avax_stargate_send": 0,
    "avax_stargate_received_usdc": 0,
    "avax_stargate_received_usdt": 0,
    "avax_stargate_stake_stg": 0,
    "avax_btc_bridge_send": 0,
    "avax_swimmer_send": 0,

    "swimmer_received": 0,

    "matic_stargate_send": 0,
    "matic_stargate_received_usdc": 0,
    "matic_stargate_received_usdt": 0,
    "matic_stargate_stake_stg": 0,
    "matic_btc_bridge_received": 0,
    "matic_angle_send": 0,
    "matic_aptos_send": 0,

    "celo_angle_received": 0,
    "celo_angle_send": 0,
    "gnosis_angle_received": 0,

    "bsc_stargate_send": 0,
    "bsc_stargate_received_usdt": 0,
    "bsc_harmony_send": 0,
    "bsc_core_send": 0,
    "bsc_core_received": 0,
    "bsc_btc_bridge_received": 0,
    "bsc_aptos_send": 0,
    "bsc_omnisea_send": 0,
    "bsc_zkbridge_send": 0,

    "moonbeam_omnisea_received": 0,
    "moonbeam_omnisea_send": 0,


    "aptos_received": 0,

    "ftm_stargate_send": 0,
    "ftm_stargate_received_usdc": 0,
    "ftm_omnisea_received": 0,

    "op_stargate_send": 0,
    "op_stargate_received_usdc": 0,

    "core_received": 0,  # tx: from addr
    "core_send": 0,  # other_addr

    "harmony_received": 0,
    "harmony_omnisea_received": 0,

    "metis_omnisea_received": 0,

    "klay_kingdom_send": 0,
    "klay_kingdom_received": 0,

    "dfk_kingdom_send": 0,
    "dfk_kingdom_received": 0,

    "ftm_abracadabra_send": 0,
    "movr_abracadabra_received": 0,  # # from addr and  token id 0x0cae51e1032e8461f4806e26332c030e34de3adb
    "movr_in": 0,

    "avax_holograph_mint": 0,
    "bsc_holograph_mint": 0,
    "matic_holograph_mint": 0,

    "avax_holograph_send": 0,
    "bsc_holograph_send": 0,
    "matic_holograph_send": 0,

    "avax_holograph_received": 0,
    "matic_holograph_received": 0,

    "opbnb_in": 0,
    "combo_in": 0
}

wallets_result = {}
project = []
for wallet in wallets:

    with open(f'data/{wallet}.json', 'r') as file:
        wallet_transaction = json.load(file)

    wallets_result[wallet] = wallet_temp.copy()
    timestamps = []
    for transaction in wallet_transaction:
        transaction["wallet"] = wallet
        project.append(transaction)

        try:

            if transaction["cate_id"] != "approve" and transaction["tx"]["status"] == 1:

                if transaction["other_addr"] == contracts["arb_stargate_send"] and transaction["chain"] == "arb":
                    wallets_result[wallet]["arb_stargate_send"] += 1
                    wallets_result[wallet]["arb_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["arb_stargate_received_usdc"] and transaction["chain"] == "arb":
                    wallets_result[wallet]["arb_stargate_received_usdc"] += 1
                    wallets_result[wallet]["arb_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == contracts["arb_stargate_received_usdt"] and transaction["chain"] == "arb":
                    wallets_result[wallet]["arb_stargate_received_usdt"] += 1
                    wallets_result[wallet]["arb_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == "0x177d36dbe2271a4ddb2ad8304d82628eb921d790" and transaction["chain"] == "arb":
                    wallets_result[wallet]["arb_stargate_received_usdt"] += 1
                    wallets_result[wallet]["arb_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == contracts["arb_testnetbridge_send"] and transaction["chain"] == "arb":
                    wallets_result[wallet]["arb_testnetbridge_send"] += 1
                    wallets_result[wallet]["testnet_in"] += 1
                    wallets_result[wallet]["testnet_bridge_received"] += 1
                    wallets_result[wallet]["arb_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["arb_stargate_stake_stg"] and transaction["chain"] == "arb":
                    wallets_result[wallet]["arb_stargate_stake_stg"] += 1

                if transaction["other_addr"] == contracts["avax_stargate_send"] and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_stargate_send"] += 1
                    wallets_result[wallet]["avax_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["avax_stargate_received_usdc"] and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_stargate_received_usdc"] += 1
                    wallets_result[wallet]["avax_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == contracts["avax_stargate_received_usdt"] and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_stargate_received_usdt"] += 1
                    wallets_result[wallet]["avax_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == "0xcd2e3622d483c7dc855f72e5eafadcd577ac78b4" and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_stargate_received_usdt"] += 1
                    wallets_result[wallet]["avax_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == contracts["avax_stargate_stake_stg"] and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_stargate_stake_stg"] += 1

                if transaction["other_addr"] == contracts["avax_btc_bridge_send"] and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_btc_bridge_send"] += 1
                    wallets_result[wallet]["avax_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["avax_swimmer_send"] and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_swimmer_send"] += 1
                    wallets_result[wallet]["avax_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["tx"]["from_addr"] == contracts["swimmer_received"] and transaction["chain"] == "swm":
                    if transaction["receives"][0]["token_id"] == "0xc1a1f40d558a3e82c3981189f61ef21e17d6eb48":
                        wallets_result[wallet]["swimmer_received"] += 1
                        wallets_result[wallet]["swm_in"] += 1

                if transaction["other_addr"] == contracts["matic_stargate_send"] and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_stargate_send"] += 1
                    wallets_result[wallet]["matic_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["matic_stargate_received_usdc"] and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_stargate_received_usdc"] += 1
                    wallets_result[wallet]["matic_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == contracts["matic_stargate_received_usdt"] and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_stargate_received_usdt"] += 1
                    wallets_result[wallet]["matic_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == "0x29e38769f23701a2e4a8ef0492e19da4604be62c" and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_stargate_received_usdt"] += 1
                    wallets_result[wallet]["matic_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == contracts["matic_stargate_stake_stg"] and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_stargate_stake_stg"] += 1

                if transaction["tx"]["from_addr"] == contracts["matic_btc_bridge_received"] and transaction["chain"] == "matic":
                    if transaction["receives"][0]["token_id"] == "0x2297aebd383787a160dd0d9f71508148769342e3":
                        wallets_result[wallet]["matic_btc_bridge_received"] += 1
                        wallets_result[wallet]["matic_in"] += 1

                if transaction["other_addr"] == contracts["matic_angle_send"] and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_angle_send"] += 1
                    wallets_result[wallet]["matic_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["celo_angle_send"] and transaction["chain"] == "celo":
                    wallets_result[wallet]["celo_angle_send"] += 1
                    wallets_result[wallet]["celo_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["matic_aptos_send"] and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_aptos_send"] += 1
                    wallets_result[wallet]["aptos_received"] += 1
                    wallets_result[wallet]["matic_from"] += 1
                    wallets_result[wallet]["aptos_in"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["tx"]["from_addr"] == contracts["celo_angle_received"] and transaction["chain"] == "celo":
                    if transaction["receives"]:
                        if transaction["receives"][0]["token_id"] == "0xc16b81af351ba9e64c1a069e3ab18c244a1e3049":
                            wallets_result[wallet]["celo_angle_received"] += 1
                            wallets_result[wallet]["celo_in"] += 1

                if transaction["tx"]["from_addr"] == contracts["gnosis_angle_received"] and transaction["chain"] == "xdai":
                    if transaction["receives"]:
                        if transaction["receives"][0]["token_id"] == "0x4b1e2c2762667331bc91648052f646d1b0d35984":
                            wallets_result[wallet]["gnosis_angle_received"] += 1
                            wallets_result[wallet]["gnosis_in"] += 1

                if transaction["other_addr"] == contracts["bsc_stargate_send"] and transaction["chain"] == "bsc":
                    wallets_result[wallet]["bsc_stargate_send"] += 1
                    wallets_result[wallet]["bsc_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["bsc_stargate_received_usdt"] and transaction["chain"] == "bsc":
                    wallets_result[wallet]["bsc_stargate_received_usdt"] += 1
                    wallets_result[wallet]["bsc_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == "0xa27a2ca24dd28ce14fb5f5844b59851f03dcf182" and transaction["chain"] == "bsc":
                    wallets_result[wallet]["bsc_stargate_received_usdt"] += 1
                    wallets_result[wallet]["bsc_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == contracts["bsc_harmony_send"] and transaction["chain"] == "bsc":
                    wallets_result[wallet]["bsc_harmony_send"] += 1
                    wallets_result[wallet]["bsc_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["bsc_core_send"] and transaction["chain"] == "bsc" and transaction["cate_id"] is None:
                    wallets_result[wallet]["bsc_core_send"] += 1
                    wallets_result[wallet]["bsc_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["bsc_core_received"] and transaction["chain"] == "bsc" and transaction["cate_id"] == "receive":
                    wallets_result[wallet]["bsc_core_received"] += 1
                    wallets_result[wallet]["bsc_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["tx"]["from_addr"] == contracts["bsc_btc_bridge_received"] and transaction["chain"] == "bsc":
                    if transaction["receives"][0]["token_id"] == "0x2297aebd383787a160dd0d9f71508148769342e3":
                        wallets_result[wallet]["bsc_btc_bridge_received"] += 1
                        wallets_result[wallet]["bsc_in"] += 1

                if transaction["other_addr"] == contracts["bsc_aptos_send"] and transaction["chain"] == "bsc":
                    wallets_result[wallet]["bsc_aptos_send"] += 1
                    wallets_result[wallet]["aptos_received"] += 1
                    wallets_result[wallet]["bsc_from"] += 1
                    wallets_result[wallet]["aptos_in"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["bsc_omnisea_send"] and transaction["chain"] == "bsc":
                    wallets_result[wallet]["bsc_omnisea_send"] += 1
                    wallets_result[wallet]["bsc_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["moonbeam_omnisea_send"] and transaction["chain"] == "mobm":
                    wallets_result[wallet]["moonbeam_omnisea_send"] += 1
                    wallets_result[wallet]["moonbeam_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["tx"]["from_addr"] == contracts["moonbeam_omnisea_received"] and transaction["chain"] == "mobm":
                    if transaction["receives"][0]["token_id"] == "0x16b26b9328b1b64e4aad6326c7cb94b2e8a96b4e":
                        wallets_result[wallet]["moonbeam_omnisea_received"] += 1
                        wallets_result[wallet]["moonbeam_in"] += 1

                if transaction["tx"]["from_addr"] == contracts["harmony_omnisea_received"] and transaction["chain"] == "hmy":
                    if transaction["receives"][0]["token_id"] == "0x20035f39c6c6cb6c0d41bb88d8f443848389b809":
                        wallets_result[wallet]["harmony_omnisea_received"] += 1
                        wallets_result[wallet]["hmy_in"] += 1

                if transaction["other_addr"] == contracts["ftm_stargate_send"] and transaction["chain"] == "ftm":
                    wallets_result[wallet]["ftm_stargate_send"] += 1
                    wallets_result[wallet]["ftm_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] == contracts["ftm_stargate_received_usdc"] and transaction["chain"] == "ftm":
                    wallets_result[wallet]["ftm_stargate_received_usdc"] += 1
                    wallets_result[wallet]["ftm_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == "0x52eea5c490fb89c7a0084b32feab854eeff07c82" and transaction["chain"] == "ftm":
                    wallets_result[wallet]["ftm_stargate_received_usdc"] += 1
                    wallets_result[wallet]["ftm_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["tx"]["from_addr"] == contracts["ftm_omnisea_received"] and transaction["chain"] == "ftm":
                    if transaction["receives"][0]["token_id"] == "0xc72633f995e98ac3bb8a89e6a9c4af335c3d6e44":
                        wallets_result[wallet]["ftm_omnisea_received"] += 1
                        wallets_result[wallet]["ftm_in"] += 1

                if transaction["other_addr"] == contracts["op_stargate_send"] and transaction["chain"] == "op":
                    wallets_result[wallet]["op_stargate_send"] += 1
                    wallets_result[wallet]["op_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["other_addr"] in [contracts["op_stargate_received_usdc1"], contracts["op_stargate_received_usdc2"]] and transaction["chain"] == "op":
                    wallets_result[wallet]["op_stargate_received_usdc"] += 1
                    wallets_result[wallet]["op_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["tx"]["from_addr"] == contracts["core_received"] and transaction["chain"] == "core":
                    wallets_result[wallet]["core_received"] += 1
                    wallets_result[wallet]["core_in"] += 1
                    wallets_result[wallet]["total_usd"] += int(transaction["receives"][0]["amount"])

                if transaction["other_addr"] == contracts["core_send"] and transaction["chain"] == "core":
                    wallets_result[wallet]["core_send"] += 1
                    wallets_result[wallet]["core_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["tx"]["from_addr"] == contracts["harmony_received"] and transaction["chain"] == "hmy":
                    wallets_result[wallet]["harmony_received"] += 1
                    wallets_result[wallet]["hmy_in"] += 1

                if transaction["tx"]["from_addr"] == contracts["metis_omnisea_received"] and transaction["chain"] == "metis":
                    if transaction["receives"][0]["token_id"] == "0xeeb51a31685bf7385b0825139320c13dce16f5fc":
                        wallets_result[wallet]["metis_omnisea_received"] += 1
                        wallets_result[wallet]["metis_in"] += 1

                # DEFI KINGDOM

                if transaction["other_addr"] == contracts["klay_kingdom_send"] and transaction["chain"] == "klay":
                    wallets_result[wallet]["klay_kingdom_send"] += 1
                    wallets_result[wallet]["klay_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["tx"]["from_addr"] == contracts["klay_kingdom_received"] and transaction["chain"] == "klay":
                    if transaction["receives"][0]["token_id"] == "0xe7a1b580942148451e47b92e95aeb8d31b0aca37":
                        wallets_result[wallet]["klay_kingdom_received"] += 1
                        wallets_result[wallet]["klay_in"] += 1

                if transaction["other_addr"] == contracts["dfk_kingdom_send"] and transaction["chain"] == "dfk":
                    wallets_result[wallet]["dfk_kingdom_send"] += 1
                    wallets_result[wallet]["dfk_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["tx"]["from_addr"] == contracts["dfk_kingdom_received"] and transaction["chain"] == "dfk":
                    if transaction["receives"][0]["token_id"] == "0x576c260513204392f0ec0bc865450872025cb1ca":
                        wallets_result[wallet]["dfk_kingdom_received"] += 1
                        wallets_result[wallet]["dfk_in"] += 1

                if transaction["other_addr"] == contracts["ftm_abracadabra_send"] and transaction["chain"] == "ftm":
                    wallets_result[wallet]["ftm_abracadabra_send"] += 1
                    wallets_result[wallet]["ftm_from"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["tx"]["from_addr"] == contracts["movr_abracadabra_received"] and transaction["chain"] == "movr":
                    if transaction["receives"]:
                        if transaction["receives"][0]["token_id"] == "0x0cae51e1032e8461f4806e26332c030e34de3adb":
                            wallets_result[wallet]["movr_abracadabra_received"] += 1
                            wallets_result[wallet]["movr_in"] += 1

                #  holograph

                if transaction["other_addr"] in ["0xe5325804d68033edf65a86403b2592a99e1f06de", "0x4803e859a2e325dc8f6adcd23ea682e323f59640"] and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_holograph_mint"] += 1

                if transaction["other_addr"] == "0xd85b5e176a30edd1915d6728faebd25669b60d8b" and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_holograph_send"] += 1
                    wallets_result[wallet]["nft_trans"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    wallets_result[wallet]["avax_from"] += 1
                    timestamps.append(int(transaction["time_at"]))


                if transaction["other_addr"] == "0xe8303f9b7d7a3de35f8c9b4405411c5ebfaf4c2c" and transaction["chain"] == "avax":
                    wallets_result[wallet]["avax_holograph_received"] += 1
                    wallets_result[wallet]["avax_in"] += 1

                if transaction["other_addr"] == "0xe5325804d68033edf65a86403b2592a99e1f06de" and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_holograph_mint"] += 1

                if transaction["other_addr"] == "0xd85b5e176a30edd1915d6728faebd25669b60d8b" and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_holograph_send"] += 1
                    wallets_result[wallet]["nft_trans"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    wallets_result[wallet]["matic_from"] += 1
                    timestamps.append(int(transaction["time_at"]))


                if transaction["other_addr"] == "0xe8303f9b7d7a3de35f8c9b4405411c5ebfaf4c2c" and transaction["chain"] == "matic":
                    wallets_result[wallet]["matic_holograph_received"] += 1
                    wallets_result[wallet]["matic_in"] += 1

                if transaction["other_addr"] == "0xe5325804d68033edf65a86403b2592a99e1f06de" and transaction["chain"] == "bsc":
                    wallets_result[wallet]["bsc_holograph_mint"] += 1

                if transaction["other_addr"] == "0xd85b5e176a30edd1915d6728faebd25669b60d8b" and transaction["chain"] == "bsc":
                    wallets_result[wallet]["bsc_holograph_send"] += 1
                    wallets_result[wallet]["nft_trans"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    wallets_result[wallet]["bsc_from"] += 1
                    timestamps.append(int(transaction["time_at"]))

                if transaction["chain"] == "bsc" and transaction["other_addr"] == "0xe09828f0da805523878be66ea2a70240d312001e" and transaction["tx"]["name"] == "transferNFT":
                    wallets_result[wallet]["bsc_zkbridge_send"] += 1
                    wallets_result[wallet]["nft_trans"] += 1
                    wallets_result[wallet]["total_trans"] += 1
                    wallets_result[wallet]["opbnb_in"] += 1
                    wallets_result[wallet]["combo_in"] += 1
                    timestamps.append(int(transaction["time_at"]))





        except Exception as e:
            print(e, "ошибка")
            print(transaction, "ошибка")

    # Подсчет положительных ключей "from" и "in"
    wallets_result[wallet]["chain_from"] = sum(
        1 for key in wallets_result[wallet] if key.endswith("_from") and wallets_result[wallet][key] > 0
    )
    wallets_result[wallet]["chain_in"] = sum(
        1 for key in wallets_result[wallet] if key.endswith("_in") and wallets_result[wallet][key] > 0
    )
    if timestamps:
        wallets_result[wallet]["last_day"] = datetime.fromtimestamp(max(timestamps)).strftime("%d.%m.%Y")
        wallets_result[wallet]["first_day"] = datetime.fromtimestamp(min(timestamps)).strftime("%d.%m.%Y")


        # print(timestamps)
        wallets_result[wallet]["days"] = len(set([datetime.fromtimestamp(ts).date() for ts in timestamps]))
        wallets_result[wallet]["weeks"] = len(set([datetime.fromtimestamp(ts).strftime('%Y-%W') for ts in timestamps]))
        wallets_result[wallet]["months"] = len(set([datetime.fromtimestamp(ts).strftime('%Y-%m') for ts in timestamps]))


    # print(wallet)
    # for title in wallets_result[wallet]:
    #     if wallets_result[wallet][title] != 0:
    #         print(title, wallets_result[wallet][title])
    # print()

headers = list(wallet_temp.keys())
headers.insert(0, "Wallet")
headers.insert(1, "chain_from")
headers.insert(2, "chain_in")

# Write headers to the first row of the sheet
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Write wallet data to the sheet
for row_num, wallet in enumerate(wallets_result.keys(), 1):
    sheet.cell(row=row_num + 1, column=1, value=wallet)  # Write wallet to the first column
    for col_num, key in enumerate(headers[1:], 2):
        sheet.cell(row=row_num + 1, column=col_num, value=wallets_result[wallet][key])  # Write value to corresponding column

# Save the workbook
workbook.save(filename)

#
for transaction in project:
    if transaction["id"] == "0x3d8fa40441a5d73615092a29674fca515668185703c75ddf43e3604e9db2277e":
        # if transaction['time_at'] < (datetime.now() - timedelta(days=7)).timestamp():
        # if transaction["id"] == "0x0ada642fd9f4876658abfcac5c5b20ad50ceb6b0e07846ab2ce7413c72a4a799":
            print("wallet:", transaction['wallet'])
            print("cate_id:", transaction['cate_id'])
            print("chain:", transaction['chain'])
            print("id:", transaction['id'])
            print("other_addr:", transaction['other_addr'])
            print("project_id:", transaction['project_id'])
            print("receives:", transaction['receives'])
            print('time_at:', transaction['time_at'])
            print("sends:", transaction['sends'])
            print("tx:", transaction['tx'])
            print()
            print(transaction)


# binance 0x161ba15a5f335c9f06bb5bbb0a9ce14076fbb645, 0x9f8c163cba728e99993abe7495f06c0a3c8ac8b9, 0x86d2660297c82ac656715e00c979fb5ca65eecc5
# okex 0x06959153b974d0d5fdfd87d561db6d8d4fa0bb0b, 0x0938c63109801ee4243a487ab84dffa2bba4589e, 0xa16f524a804beaed0d791de0aa0b5836295a2a84
# okex
