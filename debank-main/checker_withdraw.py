import json
from openpyxl import Workbook
from datetime import datetime

# Create a new Excel workbook
workbook = Workbook()

# Create a new sheet
sheet = workbook.active

# Get the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Create the filename with the current date and time
filename = f"result\withdraw_results_{current_datetime}.xlsx"


with open("wallets.txt", "r") as f:
    wallets = [row.strip() for row in f if row.strip()]

wallet_temp = {
    "transactions": 0,
    "total": 0,
}

wallets_result = {}
project = []
for wallet in wallets:

    with open(f'data/{wallet}.json', 'r') as file:
        wallet_transactions = json.load(file)

    wallets_result[wallet] = wallet_temp.copy()
    wallets_result[wallet]['transactions'] = len(wallet_transactions)
    if wallet_transactions:
        wallets_result[wallet]['total'] = wallet_transactions[0]['receives'][0]['amount']
    for transaction in wallet_transactions:
        transaction["wallet"] = wallet
        project.append(transaction)




headers = list(wallet_temp.keys())
headers.insert(0, "Wallet")

# Write headers to the first row of the sheet
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Write wallet data to the sheet
for row_num, wallet in enumerate(wallets_result.keys(), 1):
    sheet.cell(row=row_num + 1, column=1, value=wallet)  # Write wallet to the first column
    for col_num, key in enumerate(headers[1:], 2):
        sheet.cell(row=row_num + 1, column=col_num, value=wallets_result[wallet][key])  # Write value to corresponding column

# Save the workbook
workbook.save(filename)

#
# for transaction in project:
    # if transaction["cate_id"] != "":
    #     if transaction["wallet"] == "0x0022103031fc5cbe2c204e4e86286961a88f9551":
    #         print("wallet:", transaction['wallet'])
    #         print("cate_id:", transaction['cate_id'])
    #         print("chain:", transaction['chain'])
    #         print("id:", transaction['id'])
    #         print("other_addr:", transaction['other_addr'])
    #         print("project_id:", transaction['project_id'])
    #         print("receives:", transaction['receives'])
    #         print("sends:", transaction['sends'])
    #         print("tx:", transaction['tx'])
    #
    #         print(transaction)


# binance 0x161ba15a5f335c9f06bb5bbb0a9ce14076fbb645, 0x9f8c163cba728e99993abe7495f06c0a3c8ac8b9, 0x86d2660297c82ac656715e00c979fb5ca65eecc5
# okex 0x06959153b974d0d5fdfd87d561db6d8d4fa0bb0b, 0x0938c63109801ee4243a487ab84dffa2bba4589e, 0xa16f524a804beaed0d791de0aa0b5836295a2a84
# okex